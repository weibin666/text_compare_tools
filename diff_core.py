#!/usr/bin/env python3
"""
文本对比核心逻辑：可配置粒度的字符/词级 diff，输出结构化片段，
供 Excel 生成（红=删除/缺失，绿=新增）、网页预览、HTML 报告复用。

被命令行工具 text_diff.py 和 Flask 服务 app.py 共用。
"""

import html
import re
from difflib import SequenceMatcher

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Font, PatternFill, Alignment

# 字体颜色：删除/缺失=红，新增=绿，相同=黑
RED = InlineFont(color="FFFF0000")
GREEN = InlineFont(color="FF008000")
BLACK = InlineFont(color="FF000000")

# 网页/报告的颜色
_HTML_COLOR = {"del": "#d11", "ins": "#080"}

_WORD_RE = re.compile(r"\w+|\s+|[^\w\s]", re.UNICODE)
_PUNCT_RE = re.compile(r"[^\w\s]", re.UNICODE)
_WS_RE = re.compile(r"\s+")
_BAD_SHEET = re.compile(r"[\[\]:*?/\\]")

DEFAULT_PREVIEW_LIMIT = 200
DEFAULT_REPORT_LIMIT = 5000


def _tokenize(text, granularity):
    if granularity == "word":
        return _WORD_RE.findall(text)
    return list(text)


def _norm_key(tok, ignore_case, ignore_ws, ignore_punct):
    k = tok
    if ignore_case:
        k = k.lower()
    if ignore_punct:
        k = _PUNCT_RE.sub("", k)
    if ignore_ws:
        k = _WS_RE.sub("", k)
    return k


def _merge(tokens, ops):
    if not tokens:
        return []
    segs = []
    cur_text, cur_op = tokens[0], ops[0]
    for tok, op in zip(tokens[1:], ops[1:]):
        if op == cur_op:
            cur_text += tok
        else:
            segs.append((cur_text, cur_op))
            cur_text, cur_op = tok, op
    segs.append((cur_text, cur_op))
    return segs


def diff_pair(a, b, granularity="char", ignore_case=False,
              ignore_ws=False, ignore_punct=False):
    """对比两段文本，返回 (a_segments, b_segments, similarity)。"""
    a = "" if a is None else str(a)
    b = "" if b is None else str(b)
    ta, tb = _tokenize(a, granularity), _tokenize(b, granularity)
    ka = [_norm_key(t, ignore_case, ignore_ws, ignore_punct) for t in ta]
    kb = [_norm_key(t, ignore_case, ignore_ws, ignore_punct) for t in tb]

    fa = [i for i, k in enumerate(ka) if k != ""]
    fb = [j for j, k in enumerate(kb) if k != ""]
    keys_a = [ka[i] for i in fa]
    keys_b = [kb[j] for j in fb]

    op_a = ["equal"] * len(ta)
    op_b = ["equal"] * len(tb)

    sm = SequenceMatcher(None, keys_a, keys_b, autojunk=False)
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == "equal":
            continue
        if tag in ("replace", "delete"):
            for x in range(i1, i2):
                op_a[fa[x]] = "del"
        if tag in ("replace", "insert"):
            for x in range(j1, j2):
                op_b[fb[x]] = "ins"

    sim = 1.0 if (not keys_a and not keys_b) else sm.ratio()
    return _merge(ta, op_a), _merge(tb, op_b), sim


def _rich(segments):
    if not segments:
        return CellRichText("")
    blocks = []
    for text, op in segments:
        font = RED if op == "del" else GREEN if op == "ins" else BLACK
        blocks.append(TextBlock(font, text))
    return CellRichText(*blocks)


def segments_to_html(segments):
    """结构化片段 -> HTML（带颜色高亮，已转义）。"""
    parts = []
    for text, op in segments:
        esc = html.escape(text).replace("\n", "<br>")
        color = _HTML_COLOR.get(op)
        parts.append(f'<span style="color:{color};font-weight:600">{esc}</span>'
                     if color else esc)
    return "".join(parts)


def _is_diff(seg_a, seg_b):
    return (any(op != "equal" for _, op in seg_a)
            or any(op != "equal" for _, op in seg_b))


def diff_rows(pairs, *, granularity="char", ignore_case=False,
              ignore_ws=False, ignore_punct=False):
    """对所有成对文本逐行 diff，返回 row 列表（含 segments、相似度、是否差异）。"""
    opts = dict(granularity=granularity, ignore_case=ignore_case,
                ignore_ws=ignore_ws, ignore_punct=ignore_punct)
    rows = []
    for idx, (a, b) in enumerate(pairs):
        seg_a, seg_b, sim = diff_pair(a, b, **opts)
        rows.append({
            "idx": idx + 1,
            "sim": round(sim * 100, 1),
            "seg_a": seg_a,
            "seg_b": seg_b,
            "diff": _is_diff(seg_a, seg_b),
        })
    return rows


def _summary(rows):
    total = len(rows)
    diff_count = sum(1 for r in rows if r["diff"])
    avg = round(sum(r["sim"] for r in rows) / total, 1) if total else 100.0
    return total, diff_count, avg


def _safe_sheet_title(title, index, used):
    t = _BAD_SHEET.sub("_", str(title)).strip() or f"diff{index + 1}"
    t = t[:31]
    base, n = t, 1
    while t in used:
        suffix = f"_{n}"
        t = base[:31 - len(suffix)] + suffix
        n += 1
    used.add(t)
    return t


def _write_sheet(ws, rows, header_a, header_b, only_diff):
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="FFD9D9D9")
    for col, title in enumerate(["行号", header_a, header_b, "相似度", "是否不同"], start=1):
        c = ws.cell(row=1, column=col, value=title)
        c.font = header_font
        c.fill = header_fill

    wrap = Alignment(wrap_text=True, vertical="top")
    out_row = 1
    for r in rows:
        if only_diff and not r["diff"]:
            continue
        out_row += 1
        ws.cell(row=out_row, column=1, value=r["idx"])
        ws.cell(row=out_row, column=2).value = _rich(r["seg_a"])
        ws.cell(row=out_row, column=3).value = _rich(r["seg_b"])
        ws.cell(row=out_row, column=4, value=r["sim"])
        ws.cell(row=out_row, column=5, value="✗" if r["diff"] else "")
        for col in (2, 3):
            ws.cell(row=out_row, column=col).alignment = wrap

    for col, width in zip("ABCDE", (6, 50, 50, 9, 8)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"


def _preview(rows, only_diff, limit):
    out = []
    for r in rows:
        if only_diff and not r["diff"]:
            continue
        if len(out) >= limit:
            break
        out.append({
            "idx": r["idx"], "sim": r["sim"], "diff": r["diff"],
            "a_html": segments_to_html(r["seg_a"]),
            "b_html": segments_to_html(r["seg_b"]),
        })
    return out


def _section_result(rows, header_a, header_b, *, only_diff, preview_limit, title=None):
    total, diff_count, avg = _summary(rows)
    res = {
        "title": title,
        "header_a": header_a,
        "header_b": header_b,
        "total": total,
        "diff_count": diff_count,
        "avg_similarity": avg,
        "preview": _preview(rows, only_diff, preview_limit),
    }
    return res


def build_diff_workbook(pairs, header_a="文本1", header_b="文本2", *,
                        granularity="char", ignore_case=False, ignore_ws=False,
                        ignore_punct=False, only_diff=False,
                        preview_limit=DEFAULT_PREVIEW_LIMIT):
    """
    单组对比。返回 (workbook, result)，result 含统计信息与预览数据。
    （向后兼容：result 顶层仍含 total/diff_count/avg_similarity/preview。）
    """
    opts = dict(granularity=granularity, ignore_case=ignore_case,
                ignore_ws=ignore_ws, ignore_punct=ignore_punct)
    rows = diff_rows(pairs, **opts)

    wb = Workbook()
    ws = wb.active
    ws.title = "diff"
    _write_sheet(ws, rows, header_a, header_b, only_diff)

    result = _section_result(rows, header_a, header_b,
                             only_diff=only_diff, preview_limit=preview_limit)
    result["_rows"] = rows  # 供 HTML 报告复用
    return wb, result


def build_multi_diff_workbook(groups, *, granularity="char", ignore_case=False,
                              ignore_ws=False, ignore_punct=False, only_diff=False,
                              preview_limit=DEFAULT_PREVIEW_LIMIT):
    """
    多组对比（多列对比 / 多文件对比）。
    groups: [{"title","header_a","header_b","pairs"}, ...]
    返回 (workbook, sections)，sections 为每组的 result 列表。
    """
    opts = dict(granularity=granularity, ignore_case=ignore_case,
                ignore_ws=ignore_ws, ignore_punct=ignore_punct)
    wb = Workbook()
    used = set()
    sections = []
    for i, g in enumerate(groups):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = _safe_sheet_title(g.get("title") or f"diff{i + 1}", i, used)
        rows = diff_rows(g["pairs"], **opts)
        _write_sheet(ws, rows, g["header_a"], g["header_b"], only_diff)
        res = _section_result(rows, g["header_a"], g["header_b"],
                              only_diff=only_diff, preview_limit=preview_limit,
                              title=ws.title)
        res["_rows"] = rows
        sections.append(res)
    return wb, sections


# --------------------------------------------------------------------------- #
# HTML 报告（可在浏览器打开，并用「打印 / 存为 PDF」导出 PDF）
# --------------------------------------------------------------------------- #

_REPORT_CSS = """
:root { color-scheme: light; }
* { box-sizing: border-box; }
body { font-family: -apple-system, "PingFang SC", "Microsoft YaHei", sans-serif;
  color: #222; margin: 24px; }
h1 { font-size: 20px; }
h2 { font-size: 16px; margin-top: 28px; border-left: 4px solid #2563eb; padding-left: 8px; }
.meta { color: #666; font-size: 13px; margin-bottom: 8px; }
.stats { display: flex; gap: 24px; flex-wrap: wrap; margin: 8px 0 12px; }
.stats b { font-size: 18px; color: #2563eb; }
.legend { font-size: 12px; color: #666; margin-bottom: 8px; }
.legend .d { color: #d11; font-weight: 600; }
.legend .i { color: #080; font-weight: 600; }
table { width: 100%; border-collapse: collapse; font-size: 13px; }
th, td { border: 1px solid #ddd; padding: 6px 8px; vertical-align: top; text-align: left;
  word-break: break-word; }
th { background: #f3f4f6; }
td.num { text-align: center; color: #999; width: 44px; }
td.sim { text-align: center; width: 64px; }
.toolbar { margin-bottom: 16px; }
button { padding: 8px 14px; font-size: 14px; border: 1px solid #2563eb; background: #2563eb;
  color: #fff; border-radius: 6px; cursor: pointer; }
@media print { .toolbar { display: none; } body { margin: 0; } th { -webkit-print-color-adjust: exact; } }
"""


def _report_section_html(res, rows, only_diff, report_limit):
    title = res.get("title")
    head = f"<h2>{html.escape(title)}</h2>" if title else ""
    body = []
    shown = 0
    for r in rows:
        if only_diff and not r["diff"]:
            continue
        if shown >= report_limit:
            break
        shown += 1
        a_html = segments_to_html(r["seg_a"]) or '<span style="color:#bbb">（空）</span>'
        b_html = segments_to_html(r["seg_b"]) or '<span style="color:#bbb">（空）</span>'
        body.append(
            f'<tr><td class="num">{r["idx"]}</td><td>{a_html}</td>'
            f'<td>{b_html}</td><td class="sim">{r["sim"]}%</td></tr>')
    more = ""
    if res["total"] > report_limit:
        more = f'<p class="meta">报告最多显示 {report_limit} 行，完整结果请用 Excel。</p>'
    return f"""{head}
    <div class="stats">
      <div>总行数 <b>{res['total']}</b></div>
      <div>差异行 <b>{res['diff_count']}</b></div>
      <div>平均相似度 <b>{res['avg_similarity']}%</b></div>
    </div>
    {more}
    <table><thead><tr><th>行</th><th>{html.escape(res['header_a'])}</th>
    <th>{html.escape(res['header_b'])}</th><th>相似度</th></tr></thead>
    <tbody>{''.join(body)}</tbody></table>"""


def build_html_report(sections, *, title="文本对比报告", only_diff=False,
                      report_limit=DEFAULT_REPORT_LIMIT):
    """
    sections: build_diff_workbook / build_multi_diff_workbook 返回的 result(s)，
    每个 result 需含 "_rows"。返回完整的独立 HTML 字符串。
    """
    if isinstance(sections, dict):
        sections = [sections]
    parts = [_report_section_html(res, res["_rows"], only_diff, report_limit)
             for res in sections]
    return f"""<!DOCTYPE html>
<html lang="zh-CN"><head><meta charset="UTF-8">
<title>{html.escape(title)}</title><style>{_REPORT_CSS}</style></head>
<body>
<h1>{html.escape(title)}</h1>
<div class="legend"><span class="d">红 = 删除/缺失</span> ＆ <span class="i">绿 = 新增</span></div>
<div class="toolbar"><button onclick="window.print()">🖨️ 打印 / 存为 PDF</button></div>
{''.join(parts)}
</body></html>"""
