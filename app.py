#!/usr/bin/env python3
"""
文本对比 Flask 服务。

支持的上传方式：
  1. 两个 txt 文件（行一一对应）
  2. 一个 Excel / CSV 文件（指定一对或多对列名 -> 多列对比）
  3. 两个 Excel / CSV 文件（各指定一列，按行对比）
  4. 一个 jsonl 文件（每行一个 json，指定一对或多对字段名）

可配置对比粒度（字符/词）、忽略大小写/空白/标点、只导出差异行；
返回网页内高亮预览 + 统计信息 + 高亮 Excel 下载 + HTML/PDF 报告下载。
"""

import csv
import io
import json
import logging
import os
import uuid

from flask import Flask, render_template, request, send_file, jsonify, Response
from werkzeug.middleware.proxy_fix import ProxyFix
from openpyxl import load_workbook

from diff_core import build_multi_diff_workbook, build_html_report
from utils import start_cleanup_thread, RateLimiter, ext_ok

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = int(os.environ.get("MAX_UPLOAD_MB", "50")) * 1024 * 1024

# 部署在 nginx 等反代之后时，从 X-Forwarded-For 取真实客户端 IP（用于限流）
if os.environ.get("TRUST_PROXY", "0") == "1":
    app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger("text_diff")

RESULT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "results")
os.makedirs(RESULT_DIR, exist_ok=True)

RESULT_TTL = int(os.environ.get("RESULT_TTL", "1800"))
start_cleanup_thread(RESULT_DIR, ttl_seconds=RESULT_TTL)
rate_limiter = RateLimiter(max_calls=int(os.environ.get("RATE_LIMIT", "30")), window=60)

ALLOWED_EXT = {
    "txt": {"txt", "text"},
    "excel": {"xlsx", "xlsm", "csv"},
    "jsonl": {"json", "jsonl", "txt"},
}


def _get_opts(form):
    granularity = form.get("granularity", "char")
    if granularity not in ("char", "word"):
        granularity = "char"
    return dict(
        granularity=granularity,
        ignore_case=form.get("ignore_case") == "1",
        ignore_ws=form.get("ignore_ws") == "1",
        ignore_punct=form.get("ignore_punct") == "1",
        only_diff=form.get("only_diff") == "1",
    )


def _read_txt_lines(file_storage):
    text = file_storage.read().decode("utf-8-sig")
    return [line.rstrip("\r\n") for line in text.splitlines()]


def _read_table(file_storage):
    """读取 Excel 或 CSV，返回 (表头list, 数据行list)。"""
    name = file_storage.filename.lower()
    data = file_storage.read()
    if name.endswith(".csv"):
        rows = list(csv.reader(io.StringIO(data.decode("utf-8-sig"))))
        if not rows:
            raise ValueError("文件为空。")
        return [str(h).strip() for h in rows[0]], [list(r) for r in rows[1:]]
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    try:
        header = list(next(it))
    except StopIteration:
        raise ValueError("文件为空。")
    header = ["" if h is None else str(h).strip() for h in header]
    return header, [list(r) for r in it]


def _column(rows, idx):
    return [(r[idx] if idx < len(r) else "") for r in rows]


def _zip_pairs(col_a, col_b):
    n = max(len(col_a), len(col_b))
    return [(col_a[i] if i < len(col_a) else "",
             col_b[i] if i < len(col_b) else "") for i in range(n)]


def parse_txt(form, files):
    f1, f2 = files.get("txt1"), files.get("txt2")
    if not f1 or not f2 or not f1.filename or not f2.filename:
        raise ValueError("请同时上传两个 txt 文件。")
    for f in (f1, f2):
        if not ext_ok(f.filename, ALLOWED_EXT["txt"]):
            raise ValueError(f"不支持的文件类型：{f.filename}")
    pairs = _zip_pairs(_read_txt_lines(f1), _read_txt_lines(f2))
    return [{"title": "diff", "header_a": f1.filename,
             "header_b": f2.filename, "pairs": pairs}]


def _pick_pairs_cols(form):
    """读取一对或多对列名，去掉空对，返回 [(col_a, col_b), ...]。"""
    cols_a = [c.strip() for c in form.getlist("excel_col_a")]
    cols_b = [c.strip() for c in form.getlist("excel_col_b")]
    pairs = [(a, b) for a, b in zip(cols_a, cols_b) if a and b]
    if not pairs:
        raise ValueError("请填写要对比的列名（至少一对）。")
    return pairs


def parse_excel(form, files):
    f = files.get("excel")
    if not f or not f.filename:
        raise ValueError("请上传一个 Excel / CSV 文件。")
    if not ext_ok(f.filename, ALLOWED_EXT["excel"]):
        raise ValueError(f"不支持的文件类型：{f.filename}")
    col_pairs = _pick_pairs_cols(form)

    header, rows = _read_table(f)
    groups = []
    for col_a, col_b in col_pairs:
        if col_a not in header or col_b not in header:
            raise ValueError(f"列名不存在：{col_a} / {col_b}。"
                             f"文件中的列为: {', '.join(h for h in header if h)}")
        pairs = _zip_pairs(_column(rows, header.index(col_a)),
                           _column(rows, header.index(col_b)))
        groups.append({"title": f"{col_a} vs {col_b}", "header_a": col_a,
                       "header_b": col_b, "pairs": pairs})
    return groups


def parse_two_excel(form, files):
    f1, f2 = files.get("excel1"), files.get("excel2")
    if not f1 or not f2 or not f1.filename or not f2.filename:
        raise ValueError("请同时上传两个 Excel / CSV 文件。")
    for f in (f1, f2):
        if not ext_ok(f.filename, ALLOWED_EXT["excel"]):
            raise ValueError(f"不支持的文件类型：{f.filename}")
    col1 = (form.get("col1") or "").strip()
    col2 = (form.get("col2") or "").strip()
    if not col1 or not col2:
        raise ValueError("请分别填写两个文件中要对比的列名。")

    h1, r1 = _read_table(f1)
    h2, r2 = _read_table(f2)
    if col1 not in h1:
        raise ValueError(f"文件1 中没有列「{col1}」。列为: {', '.join(h for h in h1 if h)}")
    if col2 not in h2:
        raise ValueError(f"文件2 中没有列「{col2}」。列为: {', '.join(h for h in h2 if h)}")
    pairs = _zip_pairs(_column(r1, h1.index(col1)), _column(r2, h2.index(col2)))
    return [{"title": f"{f1.filename}:{col1} vs {f2.filename}:{col2}",
             "header_a": f"{f1.filename} · {col1}",
             "header_b": f"{f2.filename} · {col2}", "pairs": pairs}]


def parse_jsonl(form, files):
    f = files.get("jsonl")
    if not f or not f.filename:
        raise ValueError("请上传一个 jsonl 文件。")
    if not ext_ok(f.filename, ALLOWED_EXT["jsonl"]):
        raise ValueError(f"不支持的文件类型：{f.filename}")
    keys_a = [k.strip() for k in form.getlist("json_key_a")]
    keys_b = [k.strip() for k in form.getlist("json_key_b")]
    key_pairs = [(a, b) for a, b in zip(keys_a, keys_b) if a and b]
    if not key_pairs:
        raise ValueError("请填写要对比的字段名（至少一对，如 src / mt）。")

    objs = []
    for lineno, line in enumerate(f.read().decode("utf-8-sig").splitlines(), start=1):
        line = line.strip()
        if not line:
            continue
        try:
            objs.append(json.loads(line))
        except json.JSONDecodeError as e:
            raise ValueError(f"第 {lineno} 行不是合法 json：{e}")
    if not objs:
        raise ValueError("文件中没有有效的 json 行。")

    groups = []
    for key_a, key_b in key_pairs:
        pairs = [(o.get(key_a, ""), o.get(key_b, "")) for o in objs]
        groups.append({"title": f"{key_a} vs {key_b}", "header_a": key_a,
                       "header_b": key_b, "pairs": pairs})
    return groups


PARSERS = {"txt": parse_txt, "excel": parse_excel,
           "two_excel": parse_two_excel, "jsonl": parse_jsonl}


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/diff", methods=["POST"])
def diff():
    if not rate_limiter.allow(request.remote_addr or "?"):
        return jsonify(ok=False, error="请求过于频繁，请稍后再试。"), 429

    parser = PARSERS.get(request.form.get("mode"))
    if not parser:
        return jsonify(ok=False, error="未知的对比类型。"), 400
    try:
        groups = parser(request.form, request.files)
        opts = _get_opts(request.form)
        wb, sections = build_multi_diff_workbook(groups, **opts)

        token = uuid.uuid4().hex
        wb.save(os.path.join(RESULT_DIR, f"{token}.xlsx"))
        report_html = build_html_report(sections, only_diff=opts["only_diff"])
        with open(os.path.join(RESULT_DIR, f"{token}.html"), "w", encoding="utf-8") as fp:
            fp.write(report_html)

        total = sum(s["total"] for s in sections)
        diff_total = sum(s["diff_count"] for s in sections)
        logger.info("diff mode=%s groups=%d total=%d diff=%d",
                    request.form.get("mode"), len(sections), total, diff_total)

        # 去掉不可序列化、体积大的 _rows
        public = [{k: v for k, v in s.items() if k != "_rows"} for s in sections]
        return jsonify(
            ok=True,
            sections=public,
            download_url=f"/download/{token}",
            report_url=f"/report/{token}",
        )
    except ValueError as e:
        return jsonify(ok=False, error=str(e)), 400
    except Exception as e:  # noqa: BLE001
        logger.exception("diff failed")
        return jsonify(ok=False, error=f"处理失败：{e}"), 500


def _safe_path(token, ext):
    if not token or not all(c in "0123456789abcdef" for c in token):
        return None
    return os.path.join(RESULT_DIR, f"{token}.{ext}")


@app.route("/download/<token>")
def download(token):
    path = _safe_path(token, "xlsx")
    if path is None:
        return "非法请求", 400
    if not os.path.exists(path):
        return "文件不存在或已过期", 404
    return send_file(path, as_attachment=True, download_name="diff_result.xlsx")


@app.route("/report/<token>")
def report(token):
    path = _safe_path(token, "html")
    if path is None:
        return "非法请求", 400
    if not os.path.exists(path):
        return "文件不存在或已过期", 404
    with open(path, "r", encoding="utf-8") as fp:
        return Response(fp.read(), mimetype="text/html")


if __name__ == "__main__":
    debug = os.environ.get("FLASK_DEBUG", "0") == "1"
    app.run(host="127.0.0.1", port=5000, debug=debug)
